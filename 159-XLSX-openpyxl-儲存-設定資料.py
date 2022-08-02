# MIT License： 柯博文 老師
#  TEST 123
from openpyxl import Workbook     # pip install openpyxl
import time                       # 時間


wb = Workbook()                   # 初始化
sheet = wb.active                 # 新增一個工作欄

sheet['A1'] = 87                  # 設定資料   A1
sheet['A2'] = "Powen Ko"
sheet['A3'] = 41.80
sheet['A4'] = 10
sheet['A6'] = "柯博文老師"

now = time.strftime("%x")        # 取得 現在的時間
sheet['A5'] = now

sheet.cell(row=1, column=2).value = 'OpenPyxl Tutorial' # 設定資料 B1
sheet.cell(row=2, column=2).value =13.4                 # 設定資料  B2

wb.save("sample_file.xlsx")